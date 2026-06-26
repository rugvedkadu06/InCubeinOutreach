import os
import json
import zipfile
import io
from fastapi import FastAPI, HTTPException, Query, Response
from dotenv import load_dotenv

load_dotenv()
os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"
from fastapi.responses import StreamingResponse, JSONResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List

from .database import get_db_connection, get_pipeline_logs, log_pipeline_step, clear_all_tables
from .scraper import run_scraper_pipeline
from .cleaner import run_cleaner_pipeline
from .resolution import run_resolution_pipeline
from .enricher import run_enricher_pipeline
from .graph import (
    generate_web_graph,
    generate_csv_exports,
    generate_json_export,
    generate_mongodb_export,
    generate_neo4j_export
)

def get_region(state: str) -> str:
    if not state:
        return "Unknown"
    state_clean = state.strip().lower()
    mapping = {
        'delhi': 'North', 'haryana': 'North', 'himachal pradesh': 'North',
        'jammu and kashmir': 'North', 'jammu & kashmir': 'North',
        'punjab': 'North', 'rajasthan': 'North', 'uttarakhand': 'North', 'ladakh': 'North',
        'andhra pradesh': 'South', 'karnataka': 'South', 'kerala': 'South',
        'tamil nadu': 'South', 'telangana': 'South', 'lakshadweep': 'South',
        'puducherry': 'South', 'andaman and nicobar islands': 'South',
        'gujarat': 'West', 'maharashtra': 'West', 'goa': 'West',
        'daman & diu': 'West', 'dadra & nagar haveli': 'West',
        'bihar': 'East', 'odisha': 'East', 'west bengal': 'East', 'jharkhand': 'East',
        'chhattisgarh': 'Central', 'madhya pradesh': 'Central', 'uttar pradesh': 'Central',
        'assam': 'Northeast', 'tripura': 'Northeast', 'sikkim': 'Northeast',
        'meghalaya': 'Northeast', 'manipur': 'Northeast', 'mizoram': 'Northeast',
        'nagaland': 'Northeast', 'arunachal pradesh': 'Northeast'
    }
    return mapping.get(state_clean, "Unknown")

app = FastAPI(title="Indian Startup Ecosystem Intelligence Platform API")

# Enable CORS for frontend development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class PipelineRunResponse(BaseModel):
    status: str
    message: str
    details: Optional[dict] = None

@app.post("/api/pipeline/reset", response_model=PipelineRunResponse)
def reset_pipeline():
    try:
        clear_all_tables()
        log_pipeline_step("SYSTEM", "SUCCESS", "Ecosystem database tables reset successfully.")
        return PipelineRunResponse(status="success", message="Ecosystem database cleared.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/pipeline/run", response_model=PipelineRunResponse)
def run_pipeline(stage: str = Query("all", description="Stage to run: scrape, clean, resolve, enrich, or all")):
    try:
        stage = stage.lower()
        if stage == "scrape":
            res = run_scraper_pipeline()
            return PipelineRunResponse(status="success", message="Scraper pipeline run complete.", details=res)
        elif stage == "clean":
            res = run_cleaner_pipeline()
            return PipelineRunResponse(status="success", message="Cleaner pipeline run complete.", details=res)
        elif stage == "resolve":
            res = run_resolution_pipeline()
            return PipelineRunResponse(status="success", message="Entity resolution pipeline run complete.", details=res)
        elif stage == "enrich":
            res = run_enricher_pipeline()
            return PipelineRunResponse(status="success", message="AI enrichment pipeline run complete.", details=res)
        elif stage == "all":
            # Run all sequentially
            scrape_res = run_scraper_pipeline()
            clean_res = run_cleaner_pipeline()
            resolve_res = run_resolution_pipeline()
            enrich_res = run_enricher_pipeline()
            return PipelineRunResponse(
                status="success",
                message="Complete pipeline (Scrape -> Clean -> Resolve -> Enrich) executed successfully.",
                details={
                    "scrape": scrape_res,
                    "clean": clean_res,
                    "resolve": resolve_res,
                    "enrich": enrich_res
                }
            )
        else:
            raise HTTPException(status_code=400, detail="Invalid pipeline stage. Select from: scrape, clean, resolve, enrich, all")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/pipeline/logs")
def get_logs():
    return get_pipeline_logs()

@app.get("/api/incubators")
def get_incubators(
    q: Optional[str] = None,
    org_type: Optional[str] = None,
    state: Optional[str] = None,
    city: Optional[str] = None,
    sector: Optional[str] = None,
    region: Optional[str] = None
):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    query = "SELECT * FROM incubators WHERE 1=1"
    params = []
    
    if q:
        query += " AND (name LIKE ? OR description LIKE ? OR founder_or_head LIKE ?)"
        q_wild = f"%{q}%"
        params.extend([q_wild, q_wild, q_wild])
    if org_type:
        query += " AND organization_type = ?"
        params.append(org_type)
    if state:
        query += " AND state = ?"
        params.append(state)
    if city:
        query += " AND city = ?"
        params.append(city)
    if sector:
        query += " AND focus_areas LIKE ?"
        params.append(f"%{sector}%")
        
    query += " ORDER BY startup_count DESC"
    
    cursor.execute(query, params)
    rows = [dict(row) for row in cursor.fetchall()]
    
    # Decode JSON fields and add region mapping
    for row in rows:
        row["region"] = get_region(row["state"])
        for json_field in ["incubation_programs", "acceleration_programs", "lab_facilities", "focus_areas"]:
            if row[json_field]:
                try:
                    row[json_field] = json.loads(row[json_field])
                except:
                    row[json_field] = []
                    
    if region:
        rows = [r for r in rows if r["region"].lower() == region.lower()]
        
    conn.close()
    return rows

@app.get("/api/startups")
def get_startups(
    q: Optional[str] = None,
    sector: Optional[str] = None,
    funding_stage: Optional[str] = None,
    hq_city: Optional[str] = None,
    incubator_id: Optional[str] = None
):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    query = "SELECT * FROM startups WHERE 1=1"
    params = []
    
    if q:
        query += " AND (startup_name LIKE ? OR sector LIKE ?)"
        q_wild = f"%{q}%"
        params.extend([q_wild, q_wild])
    if sector:
        query += " AND sector = ?"
        params.append(sector)
    if funding_stage:
        query += " AND funding_stage = ?"
        params.append(funding_stage)
    if hq_city:
        query += " AND hq_city = ?"
        params.append(hq_city)
    if incubator_id:
        query += " AND incubator_id = ?"
        params.append(incubator_id)
        
    cursor.execute(query, params)
    rows = [dict(row) for row in cursor.fetchall()]
    
    # Decode JSON fields
    for row in rows:
        if row["founders"]:
            try:
                row["founders"] = json.loads(row["founders"])
            except:
                row["founders"] = []
                
    conn.close()
    return rows

@app.get("/api/graph")
def get_graph():
    return generate_web_graph()

@app.get("/api/analytics")
def get_analytics():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Simple totals
    cursor.execute("SELECT COUNT(*) FROM incubators")
    total_incubators = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT state) FROM incubators WHERE state IS NOT NULL AND state != ''")
    states_covered = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT city) FROM incubators WHERE city IS NOT NULL AND city != ''")
    cities_covered = cursor.fetchone()[0]
    
    # State-wise distribution
    cursor.execute("SELECT state, COUNT(*) as count FROM incubators WHERE state IS NOT NULL AND state != '' GROUP BY state ORDER BY count DESC")
    state_distribution = [dict(row) for row in cursor.fetchall()]
    
    # City-wise distribution (Top hubs)
    cursor.execute("SELECT city, COUNT(*) as count FROM incubators WHERE city IS NOT NULL AND city != '' GROUP BY city ORDER BY count DESC LIMIT 8")
    top_hubs = [dict(row) for row in cursor.fetchall()]
    
    # Sector distribution from focus areas of incubators
    sector_counts = {}
    cursor.execute("SELECT focus_areas FROM incubators")
    for row in cursor.fetchall():
        if row[0]:
            try:
                areas = json.loads(row[0])
                for area in areas:
                    if area:
                        sector_counts[area] = sector_counts.get(area, 0) + 1
            except:
                pass
    
    sector_distribution = [
        {"sector": k, "count": v}
        for k, v in sector_counts.items()
    ]
    sector_distribution.sort(key=lambda x: x["count"], reverse=True)
    sectors_supported = len(sector_distribution)

    # Region-wise distribution calculation
    cursor.execute("SELECT state FROM incubators")
    region_counts = {}
    for r in cursor.fetchall():
        st = r[0]
        reg = get_region(st)
        region_counts[reg] = region_counts.get(reg, 0) + 1
        
    region_distribution = [
        {"region": k, "count": v}
        for k, v in region_counts.items()
    ]
    region_distribution.sort(key=lambda x: x["count"], reverse=True)

    # Fetch unique filters for dropdowns
    cursor.execute("SELECT DISTINCT state FROM incubators WHERE state IS NOT NULL AND state != '' ORDER BY state")
    unique_states = [r[0] for r in cursor.fetchall()]
    
    cursor.execute("SELECT DISTINCT city FROM incubators WHERE city IS NOT NULL AND city != '' ORDER BY city")
    unique_cities = [r[0] for r in cursor.fetchall()]
    
    # Collect unique focus areas
    cursor.execute("SELECT focus_areas FROM incubators")
    all_areas = set()
    for row in cursor.fetchall():
        if row[0]:
            try:
                all_areas.update(json.loads(row[0]))
            except:
                pass
    unique_focus_areas = sorted(list(all_areas))

    conn.close()
    return {
        "totals": {
            "incubators": total_incubators,
            "states": states_covered,
            "cities": cities_covered,
            "sectors": sectors_supported
        },
        "state_distribution": state_distribution,
        "region_distribution": region_distribution,
        "top_hubs": top_hubs,
        "sector_distribution": sector_distribution,
        "top_incubators": [],
        "funding_stages": [],
        "filters": {
            "states": unique_states,
            "cities": unique_cities,
            "focus_areas": unique_focus_areas,
            "regions": ["North", "South", "East", "West", "Central", "Northeast"]
        }
    }

@app.get("/api/export/{format_type}")
def export_data(format_type: str):
    format_type = format_type.lower()
    
    if format_type == "json":
        data = generate_json_export()
        json_str = json.dumps(data, indent=2)
        return Response(
            content=json_str,
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=ecosystem_graph.json"}
        )
        
    elif format_type == "mongodb":
        script_str = generate_mongodb_export()
        return Response(
            content=script_str,
            media_type="text/javascript",
            headers={"Content-Disposition": "attachment; filename=import_mongodb.js"}
        )
        
    elif format_type == "neo4j":
        cypher_str = generate_neo4j_export()
        return Response(
            content=cypher_str,
            media_type="text/plain",
            headers={"Content-Disposition": "attachment; filename=import_graph.cypher"}
        )
        
    elif format_type == "csv":
        csv_files = generate_csv_exports()
        
        # Create an in-memory zip file
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for filename, csv_content in csv_files.items():
                zip_file.writestr(filename, csv_content)
                
        # Seek back to start of buffer
        zip_buffer.seek(0)
        
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=ecosystem_csvs.zip"}
        )
        
    else:
        raise HTTPException(status_code=400, detail="Invalid export format. Select from: json, csv, mongodb, neo4j")

class MouSendRequest(BaseModel):
    incubator_name: str
    incubator_email: str
    party_b_name: str
    party_b_email: str
    mou_title: str
    mou_text: str
    signature_data: str # Base64 PNG image
    recipient_email: str

@app.post("/api/mou/send")
def send_mou_email(req: MouSendRequest):
    import base64
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.image import MIMEImage
    from datetime import datetime

    # 1. Parse signature data URL
    sig_base64 = req.signature_data
    if "," in sig_base64:
        sig_base64 = sig_base64.split(",")[1]
    
    try:
        sig_bytes = base64.b64decode(sig_base64)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Base64 signature image data: {str(e)}")

    # 2. Check SMTP configuration
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = os.environ.get("SMTP_PORT", "587")
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    sender_email = os.environ.get("SENDER_EMAIL") or smtp_user

    # Try fetching Nagpur university incubator email from the database if not in env
    if not sender_email or "your_email" in sender_email:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT email FROM incubators WHERE name LIKE '%Nagpur%' OR name LIKE '%IMN%' LIMIT 1")
            row = cursor.fetchone()
            if row and row["email"]:
                sender_email = row["email"]
            conn.close()
        except Exception:
            pass
            
    if not sender_email:
        sender_email = "no-reply@rtmun.ac.in"

    # Generate HTML content
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333333; max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #dddddd; border-radius: 8px;">
        <div style="text-align: center; border-bottom: 2px solid #1e3a8a; padding-bottom: 15px; margin-bottom: 20px;">
          <h2 style="color: #1e3a8a; margin: 0; text-transform: uppercase;">Startup Ecosystem Portal</h2>
          <span style="font-size: 0.85rem; color: #666666;">Government Innovation Intelligence Platform</span>
        </div>
        
        <p>Dear Stakeholder,</p>
        <p>A new Memorandum of Understanding (MoU) has been digitally signed and executed on the Ecosystem Platform. Below is the full text of the agreement for your records.</p>
        
        <div style="background-color: #f9f9f9; padding: 25px; border-left: 4px solid #1e3a8a; margin: 20px 0; font-family: Georgia, serif; white-space: pre-wrap; font-size: 0.95rem;">
          <h3 style="text-align: center; color: #1e293b; margin-top: 0; text-transform: uppercase;">{req.mou_title}</h3>
          {req.mou_text}
        </div>
        
        <div style="margin-top: 30px; border-top: 1px dashed #cccccc; padding-top: 20px; display: table; width: 100%;">
          <div style="display: table-cell; width: 50%; text-align: center; padding: 10px;">
            <div style="font-size: 0.75rem; color: #b45309; border: 1px dashed #b45309; padding: 2px; margin-bottom: 5px; display: inline-block;">PARTY A STAMPED</div>
            <div style="width: 80%; height: 1px; background-color: #bbbbbb; margin: 40px auto 5px auto;"></div>
            <strong style="font-size: 0.85rem;">{req.incubator_name} Representative</strong><br/>
            <span style="font-size: 0.75rem; color: #888888;">Authorized Signatory</span>
          </div>
          <div style="display: table-cell; width: 50%; text-align: center; padding: 10px; vertical-align: bottom;">
            <div style="margin-bottom: 5px;">
              <img src="cid:signature_image" alt="Digital Signature" style="max-height: 50px; max-width: 150px; object-fit: contain;" />
            </div>
            <div style="width: 80%; height: 1px; background-color: #bbbbbb; margin: 5px auto 5px auto;"></div>
            <strong style="font-size: 0.85rem;">{req.party_b_name} Representative</strong><br/>
            <span style="font-size: 0.75rem; color: #888888;">Authorized Signatory</span>
          </div>
        </div>
        
        <div style="margin-top: 40px; font-size: 0.75rem; color: #999999; text-align: center; border-top: 1px solid #eeeeee; padding-top: 15px;">
          This is an automated ecosystem agreement transmission from the India Startup Ecosystem Intelligence Platform.<br/>
          Location: Ministry of Science & Technology Initiatives, Govt of India.
        </div>
      </body>
    </html>
    """

    # Register the MOU recipient as a lead in outreach_leads
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM outreach_leads WHERE email = ?", (req.recipient_email,))
        row = cursor.fetchone()
        if row:
            lead_id = row["id"]
            cursor.execute(
                "UPDATE outreach_leads SET status = 'Sent', sent_at = ?, reply_text = NULL, reply_detected_at = NULL, intent_classification = NULL, lead_score = 0, meeting_link = NULL, meeting_scheduled_at = NULL WHERE id = ?",
                (datetime.now().isoformat(), lead_id)
            )
        else:
            import uuid
            lead_id = f"lead_{uuid.uuid4().hex[:8]}"
            cursor.execute('''
                INSERT INTO outreach_leads (
                    id, incubator_id, incubator_name, email, status, sent_at, lead_score
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (lead_id, "inc_mou_" + uuid.uuid4().hex[:4], req.incubator_name, req.recipient_email, "Sent", datetime.now().isoformat(), 0))
        conn.commit()
        conn.close()
    except Exception as db_err:
        print("Error registering sent MOU as outreach lead:", db_err)

    # Helper check to see if SMTP is fully configured (not using default placeholders)
    is_smtp_ready = (
        smtp_host and smtp_user and smtp_pass and 
        "your_email" not in smtp_user and 
        "your_app_password" not in smtp_pass
    )

    if is_smtp_ready:
        # SMTP configuration is present - Send real email

        try:
            msg = MIMEMultipart("related")
            msg["Subject"] = f"Digitally Executed MOU: {req.mou_title}"
            msg["From"] = sender_email
            msg["To"] = req.recipient_email
            
            # HTML text part
            msgAlternative = MIMEMultipart("alternative")
            msg.attach(msgAlternative)
            
            msgText = MIMEText(html_content, "html")
            msgAlternative.attach(msgText)
            
            # Inline Image part
            msgImage = MIMEImage(sig_bytes, name="signature.png")
            msgImage.add_header("Content-ID", "<signature_image>")
            msgImage.add_header("Content-Disposition", "inline", filename="signature.png")
            msg.attach(msgImage)
            
            # Connect and send
            port = int(smtp_port)
            if port == 465:
                server = smtplib.SMTP_SSL(smtp_host, port, timeout=10)
            else:
                server = smtplib.SMTP(smtp_host, port, timeout=10)
                server.starttls()
                
            server.login(smtp_user, smtp_pass)
            server.sendmail(sender_email, [req.recipient_email], msg.as_string())
            server.quit()
            
            return {"status": "success", "message": f"MOU email successfully dispatched to {req.recipient_email}."}
        except Exception as smtp_err:
            raise HTTPException(status_code=500, detail=f"SMTP Server error: {str(smtp_err)}")
    else:
        # SMTP not configured - Fallback to mock log write
        try:
            scratch_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "scratch")
            os.makedirs(scratch_dir, exist_ok=True)
            
            # Write email body to txt file
            log_path = os.path.join(scratch_dir, "mou_sent_log.txt")
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(f"Timestamp: {datetime.now().isoformat()}\n")
                f.write(f"Recipient: {req.recipient_email}\n")
                f.write(f"Subject: Digitally Executed MOU: {req.mou_title}\n")
                f.write(f"Party A: {req.incubator_name} ({req.incubator_email})\n")
                f.write(f"Party B: {req.party_b_name} ({req.party_b_email})\n")
                f.write(f"MOU Title: {req.mou_title}\n")
                f.write("-" * 80 + "\n")
                f.write(req.mou_text)
                f.write("\n" + "-" * 80 + "\n")
                f.write(f"Digital Signature Image Base64 Data URL Length: {len(req.signature_data)} chars\n")
            
            # Also save signature image to disk as debug validation
            sig_img_path = os.path.join(scratch_dir, "signature_debug.png")
            with open(sig_img_path, "wb") as img_f:
                img_f.write(sig_bytes)
                
            return {
                "status": "mock_success", 
                "message": "SMTP not configured in environment variables. Email simulation successfully written to file.",
                "details": {
                    "text_log": log_path,
                    "signature_png": sig_img_path
                }
            }
        except Exception as log_err:
            raise HTTPException(status_code=500, detail=f"Failed to write mock log: {str(log_err)}")

class ContactSendRequest(BaseModel):
    incubator_name: str
    recipient_email: str
    subject: str
    message: str
    meeting_date: Optional[str] = None
    meeting_time: Optional[str] = None

@app.post("/api/contact/send")
def send_contact_email(req: ContactSendRequest):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from datetime import datetime

    # 1. Check SMTP configuration
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = os.environ.get("SMTP_PORT", "587")
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    sender_email = os.environ.get("SENDER_EMAIL") or smtp_user

    # Try fetching Nagpur university incubator email from the database if not in env
    if not sender_email or "your_email" in sender_email:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT email FROM incubators WHERE name LIKE '%Nagpur%' OR name LIKE '%IMN%' LIMIT 1")
            row = cursor.fetchone()
            if row and row["email"]:
                sender_email = row["email"]
            conn.close()
        except Exception:
            pass
            
    if not sender_email:
        sender_email = "no-reply@rtmun.ac.in"

    # Generate HTML content
    datetime_section = ""
    if req.meeting_date or req.meeting_time:
        date_str = req.meeting_date or "N/A"
        time_str = req.meeting_time or "N/A"
        datetime_section = f"""
        <div style="background-color: #fef3c7; border: 1px solid #f59e0b; border-radius: 6px; padding: 15px; margin: 20px 0; color: #92400e;">
          <h4 style="margin: 0 0 8px 0; text-transform: uppercase; font-size: 0.85rem; letter-spacing: 0.5px;">Proposed Meeting Details</h4>
          <strong>Date:</strong> {date_str}<br/>
          <strong>Time:</strong> {time_str}
        </div>
        """

    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333333; max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #dddddd; border-radius: 8px;">
        <div style="text-align: center; border-bottom: 2px solid #0891b2; padding-bottom: 15px; margin-bottom: 20px;">
          <h2 style="color: #0891b2; margin: 0; text-transform: uppercase;">Startup Ecosystem Portal</h2>
          <span style="font-size: 0.85rem; color: #666666;">Government Innovation Intelligence Platform</span>
        </div>
        
        <p>Dear {req.incubator_name} Team,</p>
        <p>You have received a new contact inquiry and meeting request via the India Startup Ecosystem Portal.</p>
        
        {datetime_section}
        
        <div style="background-color: #f9f9f9; padding: 20px; border-left: 4px solid #0891b2; margin: 20px 0; white-space: pre-wrap; font-size: 0.95rem;">
          {req.message}
        </div>
        
        <p style="font-size: 0.85rem; color: #555555;">Please respond directly to the sender or follow up via your ecosystem dashboard.</p>
        
        <div style="margin-top: 40px; font-size: 0.75rem; color: #999999; text-align: center; border-top: 1px solid #eeeeee; padding-top: 15px;">
          This is an automated ecosystem message transmission from the India Startup Ecosystem Intelligence Platform.<br/>
          Location: Ministry of Science & Technology Initiatives, Govt of India.
        </div>
      </body>
    </html>
    """

    # Helper check to see if SMTP is fully configured (not using default placeholders)
    is_smtp_ready = (
        smtp_host and smtp_user and smtp_pass and 
        "your_email" not in smtp_user and 
        "your_app_password" not in smtp_pass
    )

    if is_smtp_ready:
        # SMTP configuration is present - Send real email

        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = req.subject
            msg["From"] = sender_email
            msg["To"] = req.recipient_email
            
            msgText = MIMEText(html_content, "html")
            msg.attach(msgText)
            
            # Connect and send
            port = int(smtp_port)
            if port == 465:
                server = smtplib.SMTP_SSL(smtp_host, port, timeout=10)
            else:
                server = smtplib.SMTP(smtp_host, port, timeout=10)
                server.starttls()
                
            server.login(smtp_user, smtp_pass)
            server.sendmail(sender_email, [req.recipient_email], msg.as_string())
            server.quit()
            
            return {"status": "success", "message": f"Contact email successfully sent to {req.incubator_name} at {req.recipient_email}."}
        except Exception as smtp_err:
            raise HTTPException(status_code=500, detail=f"SMTP Server error: {str(smtp_err)}")
    else:
        # SMTP not configured - Fallback to mock log write
        try:
            scratch_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "scratch")
            os.makedirs(scratch_dir, exist_ok=True)
            
            log_path = os.path.join(scratch_dir, "contact_sent_log.txt")
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(f"Timestamp: {datetime.now().isoformat()}\n")
                f.write(f"Recipient: {req.recipient_email} ({req.incubator_name})\n")
                f.write(f"Subject: {req.subject}\n")
                if req.meeting_date or req.meeting_time:
                    f.write(f"Meeting: {req.meeting_date} at {req.meeting_time}\n")
                f.write("-" * 80 + "\n")
                f.write(req.message)
                f.write("\n" + "-" * 80 + "\n")
            
            return {
                "status": "mock_success", 
                "message": "SMTP not configured. Contact simulation successfully written to scratch file log.",
                "details": {
                    "text_log": log_path
                }
            }
        except Exception as log_err:
            raise HTTPException(status_code=500, detail=f"Failed to write mock contact log: {str(log_err)}")

class ContactUpdateRequest(BaseModel):
    id: str
    email: str
    website: str


@app.post("/api/incubators/update-contact")
def update_incubator_contact(req: ContactUpdateRequest):
    import re
    from datetime import datetime
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Fetch current details to get the canonical name of the incubator
    cursor.execute("SELECT name FROM incubators WHERE id = ?", (req.id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Incubator not found")
        
    inc_name = row["name"]
    
    # 2. Update the SQLite database
    try:
        cursor.execute(
            "UPDATE incubators SET email = ?, website = ?, last_updated = ? WHERE id = ?",
            (req.email.strip(), req.website.strip(), datetime.now().isoformat(), req.id)
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=500, detail=f"Database update failed: {str(e)}")
        
    conn.close()
    
    # 3. Update the Excel file (incubators_with_contact_details(100).xlsx or db01.xlsx)
    excel_updated = False
    excel_error = None
    try:
        import openpyxl
        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "incubators_with_contact_details(100).xlsx")
        if not os.path.exists(excel_path):
            excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "db01.xlsx")
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active
            
            rows_updated = 0
            for r in range(2, sheet.max_row + 1):
                cell_val = sheet.cell(row=r, column=1).value
                if cell_val:
                    # Check exact match or similarity
                    if str(cell_val).strip().lower() == inc_name.strip().lower():
                        sheet.cell(row=r, column=4).value = req.website.strip() # Column D (Website)
                        sheet.cell(row=r, column=5).value = req.email.strip()   # Column E (Email)
                        rows_updated += 1
            
            # If no exact match, find the best match using similarity
            if rows_updated == 0:
                best_row = None
                best_sim = 0.0
                
                # Jaccard token overlap similarity helper
                def get_words(t):
                    return {w for w in re.sub(r'[^a-zA-Z0-9\s]', ' ', t.lower()).split() if w not in {"for", "and", "the", "of", "in", "at", "on", "with", "a", "an"}}
                
                name_words = get_words(inc_name)
                for r in range(2, sheet.max_row + 1):
                    cell_val = sheet.cell(row=r, column=1).value
                    if cell_val:
                        k_words = get_words(str(cell_val))
                        if name_words and k_words:
                            sim = len(name_words.intersection(k_words)) / len(name_words.union(k_words))
                            if str(cell_val).lower() in inc_name.lower() or inc_name.lower() in str(cell_val).lower():
                                sim = max(sim, 0.6)
                            if sim > best_sim:
                                best_sim = sim
                                best_row = r
                                
                if best_row and best_sim >= 0.4:
                    sheet.cell(row=best_row, column=4).value = req.website.strip()
                    sheet.cell(row=best_row, column=5).value = req.email.strip()
                    rows_updated = 1
                    
            if rows_updated > 0:
                wb.save(excel_path)
                excel_updated = True
            else:
                excel_error = f"Could not find matching incubator row in {os.path.basename(excel_path)}"
        else:
            excel_error = f"{os.path.basename(excel_path)} file not found"
    except Exception as e:
        excel_error = f"Failed to write to {os.path.basename(excel_path)}: {str(e)}"
        
    return {
        "status": "success",
        "message": "Incubator contact details updated in SQLite.",
        "excel_status": "updated" if excel_updated else "skipped",
        "excel_error": excel_error
    }

class FinderRequest(BaseModel):
    startup_name: str
    sector: str
    hq_city: str
    stage: str
    state: Optional[str] = None

@app.post("/api/incubators/find-matches")
def find_matching_incubators(req: FinderRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM incubators")
    incubators = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    matches = []
    startup_sector = req.sector.strip().lower()
    startup_city = req.hq_city.strip().lower()
    startup_state = req.state.strip().lower() if req.state else ""
    startup_stage = req.stage.strip().lower()
    
    for inc in incubators:
        score = 0
        reasons = []
        
        # 1. Sector matching (40 points)
        focus_areas = []
        if inc["focus_areas"]:
            try:
                focus_areas = json.loads(inc["focus_areas"])
            except:
                focus_areas = []
                
        focus_areas_lower = [f.lower() for f in focus_areas]
        if startup_sector in focus_areas_lower:
            score += 40
            reasons.append(f"Direct Sector Match: Incubator specializes in {req.sector}")
        elif any(startup_sector in f or f in startup_sector for f in focus_areas_lower):
            score += 30
            reasons.append(f"Sub-Sector Match: Specializes in related areas ({', '.join(focus_areas[:3])})")
        elif not focus_areas:
            score += 15
            reasons.append("General Sector Support: Multi-sector incubator open to all verticals")
        else:
            score += 5
            
        # 2. Location matching (30 points)
        inc_city = (inc["city"] or "").strip().lower()
        inc_state = (inc["state"] or "").strip().lower()
        if startup_city and inc_city == startup_city:
            score += 30
            reasons.append(f"Local City Hub: Located in {inc['city']}, matching your startup HQ")
        elif startup_state and inc_state == startup_state:
            score += 20
            reasons.append(f"Regional State Hub: Located in {inc['city']}, {inc['state']}")
        else:
            score += 8
            reasons.append(f"National Incubator: Located in {inc['city'] or 'India'}")
            
        # 3. Funding & Scheme matching (20 points)
        grant_support = (inc["grant_support"] or "").strip().lower()
        funding_support = (inc["funding_support"] or "").strip().lower()
        
        has_grant = any(w in grant_support or w in funding_support for w in ["prayas", "birac", "big", "sisfs", "nidhi", "seed", "grant"])
        if startup_stage in ["pre-seed", "seed", "idea"]:
            if has_grant:
                score += 20
                reasons.append(f"Early-Stage Support: Offers seed grants/schemes like {inc['grant_support'] or 'NIDHI/BIRAC'}")
            else:
                score += 12
                reasons.append("Early-Stage Match: Mentorship & incubation space available")
        else:
            score += 10
            reasons.append("Growth Support: Coworking and networking infrastructure")
            
        # 4. Confidence/Capability matching (10 points)
        conf = inc["confidence_score"] or 0.90
        score += int(conf * 10)
        
        score = min(100, score)
        
        # Clean incubator JSON fields for output
        for field in ["incubation_programs", "acceleration_programs", "lab_facilities", "focus_areas"]:
            if inc[field] and isinstance(inc[field], str):
                try:
                    inc[field] = json.loads(inc[field])
                except:
                    inc[field] = []
                    
        matches.append({
            "incubator": inc,
            "match_score": score,
            "reasons": reasons
        })
        
    matches.sort(key=lambda x: x["match_score"], reverse=True)
    return matches[:5]

class ChatRequest(BaseModel):
    message: str
    model: Optional[str] = None

@app.get("/api/ai/models")
def get_ai_models():
    import requests
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        return []
    try:
        response = requests.get(
            "https://openrouter.ai/api/v1/models",
            headers={"Authorization": f"Bearer {api_key}"},
            timeout=10
        )
        if response.status_code != 200:
            return []
        models = response.json().get("data", [])
        free_models = [m["id"] for m in models if ":free" in m["id"]]
        return free_models
    except Exception as e:
        print("Error fetching OpenRouter models:", e)
        return []

@app.post("/api/ai/chat")
def chat_assistant(req: ChatRequest):
    import re
    import google.generativeai as genai
    
    query = req.message.strip()
    
    # 1. RAG Context Extraction: Match relevant records in SQLite
    conn = get_db_connection()
    cursor = conn.cursor()
    
    sectors = [
        "ai", "ml", "machine learning", "tourism", "wearables", "analytics", 
        "agritech", "agriculture", "biotech", "biotechnology", "health", 
        "healthtech", "healthcare", "medtech", "manufacturing", "agribusiness", 
        "biomedical", "engineering", "food", "renewable energy", "energy", 
        "basic science", "sciences", "iot", "chemical", "pharma", "materials", 
        "dairy", "livestock", "digital tech", "digital technologies", 
        "digital technology", "empowerment", "sustainability", "food processing", 
        "food tech", "ict", "biopharma", "it", "electronics", "industry 4.0", 
        "iomt", "circular economy", "life sciences", "robotics", "media tech", 
        "enterprise ai", "smart city", "waste processing", "assistive tech", 
        "women centric", "deeptech", "saas", "web3", "fintech", "edtech", 
        "climatetech", "spacetech", "defencetech"
    ]
    matched_sectors = [s for s in sectors if s in query.lower()]
    
    states = [
        "maharashtra", "tamil nadu", "karnataka", "delhi", "gujarat", 
        "andhra pradesh", "telangana", "punjab", "rajasthan", "uttar pradesh", 
        "west bengal", "kerala", "madhya pradesh", "chhattisgarh", "odisha", 
        "assam", "bihar", "haryana", "himachal pradesh", "jammu", "kashmir", 
        "uttarakhand", "goa", "meghalaya", "sikkim", "tripura"
    ]
    matched_states = []
    for s in states:
        if s in query.lower():
            matched_states.append(s)
            
    cities = [
        "mumbai", "chennai", "bengaluru", "bangalore", "delhi", "pune", 
        "ahmedabad", "noida", "gurgaon", "gurugram", "hyderabad", "kolkata", 
        "guntur", "visakhapatnam", "silchar", "patna", "bilaspur", "bhilai", 
        "rajkot", "solan", "srinagar", "jammu", "mysore", "indore", "kolhapur", 
        "amravati", "ludhiana", "salem", "coimbatore", "prayagraj", "kanpur", 
        "dehradun", "jaipur", "guwahati", "gautam buddha nagar", "lucknow",
        "agartala", "aligarh", "amethi", "awantipora", "bathinda", "bhilwara", 
        "bidar", "chhindwara", "cuttack", "ghaziabad", "greater noida", 
        "khurda", "kishangarh", "kottayam", "kovilpatti", "mathura", "nagpur", 
        "tumakuru", "vadlamudi", "vijayapura"
    ]
    matched_cities = []
    for c in cities:
        if c in query.lower():
            matched_cities.append(c)
            
    sql = "SELECT id, name, city, state, email, website, focus_areas, source_url FROM incubators WHERE 1=1"
    params = []
    
    if matched_states:
        state_clauses = []
        for s in matched_states:
            state_clauses.append("state LIKE ?")
            params.append(f"%{s}%")
        sql += f" AND ({' OR '.join(state_clauses)})"
        
    if matched_cities:
        city_clauses = []
        for c in matched_cities:
            if c == "bangalore":
                city_clauses.extend(["city LIKE ?", "city LIKE ?"])
                params.extend(["%bangalore%", "%bengaluru%"])
            else:
                city_clauses.append("city LIKE ?")
                params.append(f"%{c}%")
        sql += f" AND ({' OR '.join(city_clauses)})"
        
    if matched_sectors:
        sector_clauses = []
        for sec in matched_sectors:
            sector_clauses.append("focus_areas LIKE ?")
            params.append(f"%{sec}%")
        sql += f" AND ({' OR '.join(sector_clauses)})"
        
    common_short_names = ["sine", "nsrcel", "ciie", "iit", "iim", "nit", "bits", "maker village", "forge", "tihan", "kiit", "gusec", "ccmb", "c-camp", "incubimn", "imn incubation", "nagpur university"]
    matched_shorts = [sn for sn in common_short_names if sn in query.lower()]
    if matched_shorts:
        name_clauses = []
        for ns in matched_shorts:
            name_clauses.append("name LIKE ?")
            params.append(f"%{ns}%")
        sql += f" AND ({' OR '.join(name_clauses)})"
        
    sql += " LIMIT 15"
    
    cursor.execute(sql, params)
    rows = [dict(r) for r in cursor.fetchall()]
    
    if not rows:
        cursor.execute("SELECT id, name, city, state, email, website, focus_areas, source_url FROM incubators ORDER BY name LIMIT 10")
        rows = [dict(r) for r in cursor.fetchall()]
        
    conn.close()
    
    context_items = []
    for r in rows:
        focus = ""
        if r["focus_areas"]:
            try:
                focus = ", ".join(json.loads(r["focus_areas"]))
            except:
                focus = r["focus_areas"]
        item = f"- Name: {r['name']}\n  Location: {r['city']}, {r['state']}\n  Sectors: {focus}\n  Website: {r['website'] or 'N/A'}\n  Email: {r['email'] or 'N/A'}\n  Funding Source: {r['source_url'] or 'N/A'}"
        context_items.append(item)
        
    context_str = "\n\n".join(context_items)
    
    openrouter_key = os.environ.get("OPENROUTER_API_KEY")
    if openrouter_key:
        import requests
        import random
        import time
        
        # 1. Fetch available free models
        free_models = []
        try:
            models_response = requests.get(
                "https://openrouter.ai/api/v1/models",
                headers={"Authorization": f"Bearer {openrouter_key}"},
                timeout=10
            )
            if models_response.status_code == 200:
                models_data = models_response.json().get("data", [])
                free_models = [m["id"] for m in models_data if ":free" in m["id"]]
        except Exception as e:
            print("Failed to fetch OpenRouter free models:", e)
            
        if free_models:
            # Prioritize selected model
            selected_model = req.model
            if selected_model and selected_model in free_models:
                free_models.remove(selected_model)
                free_models.insert(0, selected_model)
            else:
                # Default preference models
                pref_models = [m for m in free_models if "llama-3" in m or "mistral" in m]
                for pm in reversed(pref_models):
                    if pm in free_models:
                        free_models.remove(pm)
                        free_models.insert(0, pm)
            
            # Shuffle remaining models
            first = free_models[0]
            rest = free_models[1:]
            random.shuffle(rest)
            models_to_try = [first] + rest
            
            system_prompt = f"""
You are an intelligent assistant for the Indian Startup Ecosystem Intelligence Portal.

Follow these rules:

1. If the user asks about a technical term, concept, abbreviation, or definition:
   → Respond in the structured format below.

2. If the user sends a greeting (hi, hello, hey, etc.):
   → Respond normally in a friendly conversational way.

3. If the user asks a normal question:
   → Answer clearly but do NOT force the structured template.

Structured format (ONLY when explaining a term):

You likely meant **<term>**, which stands for **<full form if applicable>**.

Here's a clear explanation:

### What is <term>?
<definition>

### Key Characteristics:
1. <point>
2. <point>
3. <point>

### Examples:
- <example>
- <example>

### How It Works:
<explanation>

### Limitations:
- <limitation>
- <limitation>

Use the following context from our database to formulate your response. Suggest specific incubators by name, location, and focus areas where relevant. If the context is not directly matching, answer the question generally using your knowledge but state that these are general facts and list the closest matches from the database context.

Database Context:
{context_str}
"""
            
            for model in models_to_try:
                try:
                    start = time.time()
                    response = requests.post(
                        "https://openrouter.ai/api/v1/chat/completions",
                        headers={
                            "Authorization": f"Bearer {openrouter_key}",
                            "Content-Type": "application/json"
                        },
                        json={
                            "model": model,
                            "messages": [
                                {"role": "system", "content": system_prompt},
                                {"role": "user", "content": query}
                            ],
                            "temperature": 0.3
                        },
                        timeout=12
                    )
                    
                    if response.status_code == 200:
                        result = response.json()
                        reply = result["choices"][0]["message"]["content"]
                        response_time = round(time.time() - start, 2)
                        
                        return {
                            "status": "success",
                            "mode": "openrouter",
                            "message": reply,
                            "model_used": model,
                            "response_time": response_time
                        }
                except Exception as e:
                    print(f"❌ {model} failed:", e)
                    continue

    # Fallback 2: Gemini API
    api_key = os.environ.get("GEMINI_API_KEY")
    if api_key:
        try:
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel('gemini-1.5-flash')
            
            prompt = f"""
            You are the AI Ecosystem Assistant for the Indian Startup Ecosystem Intelligence Portal. 
            Answer the user's question clearly, professionally, and concisely using markdown formatting.
            
            User Question: "{query}"
            
            Use the following context from our database to formulate your response. Suggest specific incubators by name, location, and focus areas where relevant. If the context is not directly matching, answer the question generally using your knowledge but state that these are general facts and list the closest matches from the database.
            
            Database Context:
            {context_str}
            
            Guidelines:
            - Format answers with lists, bold text, and tables where applicable to make them highly readable.
            - Ensure any links or emails mentioned are formatted as clickable markdown links.
            - Keep response under 3 paragraphs if possible.
            """
            
            response = model.generate_content(prompt)
            return {
                "status": "success",
                "mode": "gemini_ai",
                "message": response.text.strip(),
                "model_used": "gemini-1.5-flash",
                "response_time": 1.0
            }
        except Exception as e:
            pass
            
    # Fallback 3: Local database keyword search summary
    markdown_result = f"### Ecosystem Assistant (Local Database Match Mode)\n\n"
    markdown_result += f"The AI Assistant is running in local search mode. Based on your query keywords, here are the most relevant incubator records found in the platform database:\n\n"
    markdown_result += "| Incubator Name | Location | Focus Sectors | Website |\n"
    markdown_result += "| :--- | :--- | :--- | :--- |\n"
    
    for r in rows:
        focus = "General"
        if r["focus_areas"]:
            try:
                focus = ", ".join(json.loads(r["focus_areas"])[:3])
            except:
                focus = r["focus_areas"]
        loc = f"{r['city'] or ''}, {r['state'] or ''}".strip(", ")
        markdown_result += f"| **{r['name']}** | {loc or 'India'} | {focus} | {r['website'] or 'N/A'} |\n"
        
    markdown_result += "\n\n**Suggestions for next steps:**\n"
    markdown_result += "- Click on the **Incubators Directory** tab to search and filter these records in detail.\n"
    markdown_result += "- Double-check spelling of states or cities in your question for better query matching.\n"
    
    return {
        "status": "success",
        "mode": "local_fallback",
        "message": markdown_result
    }

# Pydantic models for outreach
class OutreachEmailRequest(BaseModel):
    lead_id: str

class OutreachReplyRequest(BaseModel):
    lead_id: str
    reply_text: str

class AddLeadRequest(BaseModel):
    incubator_id: str
    incubator_name: str
    email: str

class UpdateMeetingStatusRequest(BaseModel):
    meeting_id: str
    status: str

class UpdateLeadStatusRequest(BaseModel):
    lead_id: str
    status: str

class UpdateLeadNotesRequest(BaseModel):
    lead_id: str
    notes: str

def seed_outreach_leads():
    import uuid
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Ensure trial incubator exists in incubators table
    cursor.execute("SELECT COUNT(*) FROM incubators WHERE id = 'inc_trial_rugved'")
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO incubators (
                id, name, email, city, state, organization_type, confidence_score, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            "inc_trial_rugved",
            "Trial Incubator (kadurugved0)",
            "kadurugved0@gmail.com",
            "Nagpur",
            "Maharashtra",
            "Academic Collab Partner",
            1.0,
            "resolved"
        ))
        conn.commit()

    # Ensure trial lead exists in outreach_leads
    cursor.execute("SELECT COUNT(*) FROM outreach_leads WHERE email = 'kadurugved0@gmail.com'")
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO outreach_leads (
                id, incubator_id, incubator_name, email, status, lead_score
            ) VALUES (?, ?, ?, ?, ?, ?)
        ''', ("lead_trial_rugved", "inc_trial_rugved", "Trial Incubator (kadurugved0)", "kadurugved0@gmail.com", "Draft", 0))
        conn.commit()

    # Check if leads exist
    cursor.execute("SELECT COUNT(*) FROM outreach_leads")
    count = cursor.fetchone()[0]
    if count > 1:
        conn.close()
        return
        
    # Fetch some incubators (prioritizing Nagpur University / IncubIMN)
    cursor.execute("SELECT id, name, email FROM incubators WHERE id != 'inc_trial_rudveg' ORDER BY CASE WHEN name LIKE '%IncubIMN%' THEN 0 ELSE 1 END, name LIMIT 5")
    incubators = cursor.fetchall()
    
    for inc in incubators:
        inc_id = inc[0]
        inc_name = inc[1]
        inc_email = inc[2] or f"contact@{inc_name.lower().replace(' ', '').replace(',', '').replace('(', '').replace(')', '')}.org"
        
        cursor.execute("SELECT COUNT(*) FROM outreach_leads WHERE email = ?", (inc_email,))
        if cursor.fetchone()[0] == 0:
            lead_id = f"lead_{uuid.uuid4().hex[:8]}"
            cursor.execute('''
                INSERT INTO outreach_leads (
                    id, incubator_id, incubator_name, email, status, lead_score
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (lead_id, inc_id, inc_name, inc_email, "Draft", 0))
        
    conn.commit()
    conn.close()

@app.get("/api/outreach/leads")
def get_outreach_leads():
    seed_outreach_leads()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM outreach_leads")
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

@app.post("/api/outreach/add-lead")
def add_outreach_lead(req: AddLeadRequest):
    import uuid
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if already exists in campaign
    cursor.execute("SELECT COUNT(*) FROM outreach_leads WHERE incubator_id = ? OR email = ?", (req.incubator_id, req.email))
    if cursor.fetchone()[0] > 0:
        conn.close()
        return {"status": "exists", "message": "Lead already exists in campaign leads."}
        
    lead_id = f"lead_{uuid.uuid4().hex[:8]}"
    cursor.execute('''
        INSERT INTO outreach_leads (id, incubator_id, incubator_name, email, status, lead_score)
        VALUES (?, ?, ?, ?, 'Draft', 0)
    ''', (lead_id, req.incubator_id, req.incubator_name, req.email))
    
    conn.commit()
    conn.close()
    return {"status": "success", "message": f"Successfully added {req.incubator_name} to campaigns.", "lead_id": lead_id}

@app.post("/api/outreach/reset")
def reset_outreach():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM scheduled_meetings")
    cursor.execute("DELETE FROM outreach_leads")
    conn.commit()
    conn.close()
    seed_outreach_leads()
    return {"status": "success", "message": "Campaign data successfully reset."}

@app.post("/api/outreach/leads/update-status")
def update_lead_status(req: UpdateLeadStatusRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM outreach_leads WHERE id = ?", (req.lead_id,))
    lead = cursor.fetchone()
    if not lead:
        conn.close()
        raise HTTPException(status_code=404, detail="Lead not found")
    cursor.execute("UPDATE outreach_leads SET status = ? WHERE id = ?", (req.status, req.lead_id))
    conn.commit()
    conn.close()
    return {"status": "success", "message": f"Campaign lead status updated to {req.status}."}

@app.post("/api/outreach/leads/update-notes")
def update_lead_notes(req: UpdateLeadNotesRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM outreach_leads WHERE id = ?", (req.lead_id,))
    lead = cursor.fetchone()
    if not lead:
        conn.close()
        raise HTTPException(status_code=404, detail="Lead not found")
    cursor.execute("UPDATE outreach_leads SET notes = ? WHERE id = ?", (req.notes, req.lead_id))
    conn.commit()
    conn.close()
    return {"status": "success", "message": "Campaign lead notes updated successfully."}

@app.post("/api/outreach/send-email")
def trigger_outreach_email(req: OutreachEmailRequest):
    from datetime import datetime
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM outreach_leads WHERE id = ?", (req.lead_id,))
    lead = cursor.fetchone()
    if not lead:
        conn.close()
        raise HTTPException(status_code=404, detail="Lead not found")
        
    # Send a real outreach invite email if SMTP is configured
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = os.environ.get("SMTP_PORT", "587")
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    sender_email = os.environ.get("SENDER_EMAIL") or smtp_user or "no-reply@rtmun.ac.in"
    
    is_smtp_ready = smtp_host and smtp_user and smtp_pass and "your_email" not in smtp_user
    
    email_sent_successfully = False
    if is_smtp_ready:
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            
            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"Academic Partnership Opportunity - RTMUN Innovation Ecosystem"
            msg["From"] = sender_email
            msg["To"] = lead["email"]
            
            body_text = f"""
Dear Representative,

We hope this email finds you well. 

We are reaching out from the RTMUN Startup Platform regarding a potential Strategic Cooperation and Academic Collaboration. We would love to share a draft MoU agreement with your incubation center and explore mutually beneficial synergies.

Please reply to this email to confirm your interest and schedule an introductory virtual meeting.

Sincerely,
RTMUN Innovation Portal Admin
            """
            msg.attach(MIMEText(body_text, "plain"))
            
            port = int(smtp_port)
            if port == 465:
                server = smtplib.SMTP_SSL(smtp_host, port, timeout=10)
            else:
                server = smtplib.SMTP(smtp_host, port, timeout=10)
                server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(sender_email, [lead["email"]], msg.as_string())
            server.quit()
            email_sent_successfully = True
        except Exception as e:
            print("Error sending outreach email via SMTP:", e)
            
    cursor.execute(
        "UPDATE outreach_leads SET status = 'Sent', sent_at = ? WHERE id = ?",
        (datetime.now().isoformat(), req.lead_id)
    )
    conn.commit()
    conn.close()
    msg_status = "Real email sent via SMTP" if email_sent_successfully else "SMTP not configured, simulated sending"
    return {"status": "success", "message": f"Outreach email campaign successfully triggered for {lead['incubator_name']} ({msg_status})."}

def process_reply(lead_id: str, reply_text: str):
    import uuid
    import json
    import os
    from datetime import datetime
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM outreach_leads WHERE id = ?", (lead_id,))
    lead = cursor.fetchone()
    if not lead:
        conn.close()
        return None
        
    api_key = os.environ.get("GEMINI_API_KEY")
    intent = "Neutral"
    score = 50
    summary = "Acknowledge receipt and waiting for more context."
    
    if api_key and "your_gemini" not in api_key:
        try:
            import google.generativeai as genai
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel('gemini-1.5-flash')
            
            prompt = f"""
            You are the AI Intent Classifier for the Indian Startup Ecosystem Outreach System.
            Analyze the following email reply from an incubator and output a JSON object containing:
            1. "intent": One of "Positive", "Neutral", "Negative", "Information Request"
            2. "score": An interest score from 0 to 100 representing how eager they are to collaborate/arrange a meeting. (e.g. "we would love to meet" = 90+, "tell me more" = 60-79, "no thanks" = <30).
            3. "summary": A 1-sentence summary of their response.
            
            Reply text: "{reply_text}"
            
            Output ONLY valid JSON.
            """
            response = model.generate_content(prompt)
            result_text = response.text.strip()
            if "```json" in result_text:
                result_text = result_text.split("```json")[1].split("```")[0].strip()
            elif "```" in result_text:
                result_text = result_text.split("```")[1].split("```")[0].strip()
                
            res = json.loads(result_text)
            intent = res.get("intent", intent)
            score = int(res.get("score", score))
            summary = res.get("summary", summary)
        except Exception as e:
            print("Gemini analysis error:", e)
            
    # Local fallback rules
    if not api_key or "your_gemini" in api_key or score == 50:
        reply_lower = reply_text.lower()
        if any(w in reply_lower for w in ["love to", "excited", "schedule", "meeting", "meet", "calendar", "join", "interested", "sure", "definitely", "yes", "great", "mou"]):
            intent = "Positive"
            score = 90
            summary = "Expressed strong interest and requested to schedule a meeting."
        elif any(w in reply_lower for w in ["no", "decline", "not interested", "sorry", "cannot", "cancel", "busy", "unable"]):
            intent = "Negative"
            score = 15
            summary = "Declined the collaboration proposal."
        elif any(w in reply_lower for w in ["what", "how", "information", "details", "docs", "document", "questions", "send me"]):
            intent = "Information Request"
            score = 70
            summary = "Requested additional information or documentation."
            
    meeting_details = None
    meeting_link = None
    meeting_scheduled_at = None
    
    if score < 30:
        new_status = 'Not Interested'
    else:
        new_status = 'Replied'
        
    cursor.execute('''
        UPDATE outreach_leads 
        SET status = ?, 
            reply_text = ?, 
            reply_detected_at = ?, 
            intent_classification = ?, 
            lead_score = ?,
            meeting_link = ?,
            meeting_scheduled_at = ?
        WHERE id = ?
    ''', (
        new_status,
        reply_text,
        datetime.now().isoformat(),
        intent,
        score,
        meeting_link if meeting_link else lead["meeting_link"],
        meeting_scheduled_at if meeting_scheduled_at else lead["meeting_scheduled_at"],
        lead_id
    ))
    
    conn.commit()
    conn.close()
    
    return {
        "status": "success",
        "lead_status": new_status,
        "intent": intent,
        "score": score,
        "summary": summary,
        "meeting": meeting_details
    }

@app.post("/api/outreach/simulate-reply")
def process_outreach_reply(req: OutreachReplyRequest):
    res = process_reply(req.lead_id, req.reply_text)
    if not res:
        raise HTTPException(status_code=404, detail="Lead not found")
    return res

def extract_email_address(from_header):
    if not from_header:
        return ""
    import re
    match = re.search(r'<([^>]+)>', from_header)
    if match:
        return match.group(1).strip().lower()
    # Remove surrounding quotes if any
    return from_header.strip().strip('"').strip("'").lower()

def check_imap_replies_sync():
    import imaplib
    import email
    from email.header import decode_header
    import os
    
    imap_host = os.environ.get("IMAP_HOST")
    imap_port = os.environ.get("IMAP_PORT", "993")
    imap_user = os.environ.get("IMAP_USER")
    imap_pass = os.environ.get("IMAP_PASS")
    
    if imap_user:
        imap_user = imap_user.strip().strip('"').strip("'")
    if imap_pass:
        imap_pass = imap_pass.strip().strip('"').strip("'")
        
    if not (imap_host and imap_user and imap_pass) or "your_email" in imap_user:
        print("IMAP parameters not configured or using placeholders. Skipping IMAP scan.")
        return []
        
    processed_replies = []
    try:
        print(f"Connecting to IMAP server {imap_host}:{imap_port} for user {imap_user}...")
        mail = imaplib.IMAP4_SSL(imap_host, int(imap_port))
        mail.login(imap_user, imap_pass)
        mail.select("inbox")
        
        # Search ALL messages in the inbox (Seen and Unseen)
        status, response = mail.search(None, "ALL")
        if status != "OK":
            mail.logout()
            print("IMAP search failed.")
            return []
            
        mail_ids = response[0].split()
        if not mail_ids:
            mail.logout()
            print("Inbox is empty.")
            return []
            
        # Inspect the last 40 messages to find replies
        mail_ids = mail_ids[-40:]
        print(f"IMAP connection successful. Scanning the last {len(mail_ids)} messages in Inbox...")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get active leads email map that are currently waiting for a reply
        cursor.execute("SELECT id, email, status, incubator_name FROM outreach_leads WHERE status = 'Sent'")
        leads = {row["email"].lower().strip(): row for row in cursor.fetchall()}
        
        print("Active leads waiting for replies (Sent status):", list(leads.keys()))
        
        for m_id in reversed(mail_ids): # Scan from newest to oldest
            if len(leads) == 0:
                print("All pending outreach replies have been detected and processed. Stopping search early.")
                break
                
            try:
                status, msg_data = mail.fetch(m_id, "(RFC822)")
                if status != "OK":
                    continue
                    
                raw_email = msg_data[0][1]
                msg = email.message_from_bytes(raw_email)
                
                from_header = msg.get("From", "")
                from_email = extract_email_address(from_header)
                
                print(f"Checking message From: {from_header} (parsed: {from_email})")
                
                if from_email in leads:
                    lead = leads[from_email]
                    print(f"Match found for Sent lead: {lead['incubator_name']} ({from_email})!")
                    
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            content_type = part.get_content_type()
                            content_disposition = str(part.get("Content-Disposition"))
                            if content_type == "text/plain" and "attachment" not in content_disposition:
                                payload = part.get_payload(decode=True)
                                body = payload.decode(part.get_content_charset() or "utf-8", errors="ignore")
                                break
                    else:
                        payload = msg.get_payload(decode=True)
                        body = payload.decode(msg.get_content_charset() or "utf-8", errors="ignore")
                    
                    res = process_reply(lead["id"], body)
                    if res:
                        print(f"Successfully processed reply from {from_email}. New status: {res['lead_status']}")
                        processed_replies.append({
                            "incubator_name": lead["incubator_name"],
                            "email": from_email,
                            "reply_text": body,
                            "intent": res["intent"],
                            "score": res["score"],
                            "status": res["lead_status"]
                        })
                        
                        # Remove from the temporary leads dictionary so we don't process older emails from the same sender in this run
                        del leads[from_email]
                    
                    # Mark Seen (optional since it was already scanned, but keeps inbox clean)
                    mail.store(m_id, "+FLAGS", "\\Seen")
            except Exception as item_err:
                print(f"Error reading message ID {m_id}: {item_err}")
                
        conn.close()
        mail.close()
        mail.logout()
    except Exception as e:
        print("IMAP sync exception:", e)
        
    return processed_replies

def get_google_client_config():
    client_id = os.environ.get("GOOGLE_CLIENT_ID")
    client_secret = os.environ.get("GOOGLE_CLIENT_SECRET")
    redirect_uri = os.environ.get("GOOGLE_REDIRECT_URI")
    
    if not client_id or not client_secret or "your_google_client_id" in client_id:
        return None
        
    return {
        "web": {
            "client_id": client_id.strip().strip('"').strip("'"),
            "client_secret": client_secret.strip().strip('"').strip("'"),
            "redirect_uris": [redirect_uri.strip().strip('"').strip("'")],
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token"
        }
    }

def create_gmeet_event(summary: str, description: str, date_str: str, time_str: str, attendee_email: str):
    import os
    import uuid
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from datetime import datetime, timedelta
    
    token_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "token.json")
    if not os.path.exists(token_path):
        print(f"Token file not found at {token_path}. Skipping API event scheduling.")
        return None
        
    try:
        credentials = Credentials.from_authorized_user_file(token_path)
        service = build("calendar", "v3", credentials=credentials)
        
        try:
            start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %I:%M %p")
        except Exception:
            try:
                start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
            except Exception:
                start_dt = datetime.now() + timedelta(days=2)
                
        end_dt = start_dt + timedelta(minutes=30)
        start_iso = start_dt.isoformat()
        end_iso = end_dt.isoformat()
        
        event = {
            'summary': summary,
            'description': description,
            'start': {
                'dateTime': start_iso,
                'timeZone': 'Asia/Kolkata',
            },
            'end': {
                'dateTime': end_iso,
                'timeZone': 'Asia/Kolkata',
            },
            'attendees': [
                {'email': attendee_email},
            ],
            'conferenceData': {
                'createRequest': {
                    'requestId': f"meet_{uuid.uuid4().hex[:12]}",
                    'conferenceSolutionKey': {
                        'type': 'hangoutsMeet'
                    }
                }
            }
        }
        
        created_event = service.events().insert(
            calendarId='primary',
            body=event,
            conferenceDataVersion=1
        ).execute()
        
        meet_link = None
        conf_data = created_event.get("conferenceData", {})
        entry_points = conf_data.get("entryPoints", [])
        for ep in entry_points:
            if ep.get("entryPointType") == "video":
                meet_link = ep.get("uri")
                break
                
        if not meet_link:
            meet_link = created_event.get("htmlLink")
            
        return {
            "calendar_event_id": created_event.get("id"),
            "meet_link": meet_link
        }
    except Exception as e:
        print("Error creating Google Meet calendar event:", e)
        return None

def generate_ics_file_content(summary: str, description: str, date_str: str, time_str: str, meet_link: str, organizer_email: str, attendee_email: str) -> str:
    from datetime import datetime, timedelta
    try:
        start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %I:%M %p")
    except Exception:
        try:
            start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
        except Exception:
            start_dt = datetime.now() + timedelta(days=2)
            
    end_dt = start_dt + timedelta(minutes=30)
    
    # Calculate offset for UTC
    start_utc = start_dt - timedelta(hours=5, minutes=30)
    end_utc = end_dt - timedelta(hours=5, minutes=30)
    
    start_str = start_utc.strftime("%Y%m%dT%H%M%SZ")
    end_str = end_utc.strftime("%Y%m%dT%H%M%SZ")
    dtstamp_str = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    
    uid = f"meet-{start_str}-{attendee_email.replace('@', '-')}"
    
    esc_summary = summary.replace(",", "\\,").replace(";", "\\;")
    esc_description = description.replace(",", "\\,").replace(";", "\\;").replace("\n", "\\n")
    
    ics_lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//RTMUN Startup Platform//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:REQUEST",
        "BEGIN:VEVENT",
        f"DTSTART:{start_str}",
        f"DTEND:{end_str}",
        f"DTSTAMP:{dtstamp_str}",
        f"ORGANIZER;CN=RTMUN Admin:mailto:{organizer_email}",
        f"ATTENDEE;CUTYPE=INDIVIDUAL;ROLE=REQ-PARTICIPANT;PARTSTAT=NEEDS-ACTION;RSVP=TRUE;CN={attendee_email}:mailto:{attendee_email}",
        f"UID:{uid}",
        f"SUMMARY:{esc_summary}",
        f"DESCRIPTION:{esc_description}\\n\\nJoin Video Call: {meet_link}",
        f"LOCATION:{meet_link}",
        "STATUS:CONFIRMED",
        "SEQUENCE:0",
        "TRANSP:OPAQUE",
        "BEGIN:VALARM",
        "TRIGGER:-PT15M",
        "ACTION:DISPLAY",
        "DESCRIPTION:Reminder",
        "END:VALARM",
        "END:VEVENT",
        "END:VCALENDAR"
    ]
    return "\r\n".join(ics_lines)

def send_meeting_invite_email(lead_name: str, lead_email: str, date_str: str, time_str: str, meet_link: str):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    import os
    
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = os.environ.get("SMTP_PORT", "587")
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    sender_email = os.environ.get("SENDER_EMAIL") or smtp_user or "no-reply@rtmun.ac.in"
    
    if smtp_user:
        smtp_user = smtp_user.strip().strip('"').strip("'")
    if smtp_pass:
        smtp_pass = smtp_pass.strip().strip('"').strip("'")
    if sender_email:
        sender_email = sender_email.strip().strip('"').strip("'")
        
    is_smtp_ready = smtp_host and smtp_user and smtp_pass and "your_email" not in smtp_user
    if not is_smtp_ready:
        print("SMTP credentials are not configured or using placeholders. Skipping email invitation delivery.")
        return False
        
    try:
        msg = MIMEMultipart("mixed")
        msg["Subject"] = f"Invitation: MOU Collaboration Discussion @ {date_str} {time_str}"
        msg["From"] = f"RTMUN Outreach <{sender_email}>"
        msg["To"] = lead_email
        
        summary = f"MOU Collaboration: {lead_name}"
        description = f"Dear Team at {lead_name},\n\nWe have scheduled a virtual introductory meeting to discuss our Strategic Cooperation and Academic Collaboration MOU.\n\nMeeting link: {meet_link}\n\nWe look forward to meeting you."
        
        msg_alternative = MIMEMultipart("alternative")
        msg.attach(msg_alternative)
        
        body_text = f"{description}\n\nGoogle Meet: {meet_link}\nDate: {date_str}\nTime: {time_str}"
        msg_alternative.attach(MIMEText(body_text, "plain"))
        
        body_html = f"""
        <html>
            <body style="font-family: sans-serif; line-height: 1.5; color: #1e293b;">
                <div style="max-width: 600px; margin: 0 auto; padding: 1.5rem; border: 1px solid #e2e8f0; border-radius: 8px;">
                    <h2 style="color: #8b5cf6; margin-top: 0;">MOU Collaboration Partnership Discussion</h2>
                    <p>Dear Team at <strong>{lead_name}</strong>,</p>
                    <p>We are pleased to invite you to a virtual meeting to discuss the strategic collaboration MOU between our institutions.</p>
                    
                    <div style="background: #f8fafc; padding: 1rem; border-radius: 6px; margin: 1.5rem 0;">
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr>
                                <td style="width: 80px; padding: 4px 0; color: #64748b; font-weight: 600;">Date:</td>
                                <td style="padding: 4px 0; font-weight: 700;">{date_str}</td>
                            </tr>
                            <tr>
                                <td style="color: #64748b; font-weight: 600; padding: 4px 0;">Time:</td>
                                <td style="padding: 4px 0; font-weight: 700;">{time_str}</td>
                            </tr>
                            <tr>
                                <td style="color: #64748b; font-weight: 600; padding: 4px 0;">Video Call:</td>
                                <td style="padding: 4px 0;">
                                    <a href="{meet_link}" style="color: #06b6d4; text-decoration: none; font-weight: 700;">Join Google Meet</a>
                                </td>
                            </tr>
                        </table>
                    </div>
                    
                    <p>An interactive calendar invite has been attached to this email. You can RSVP directly using your email client's RSVP/buttons.</p>
                    <p>Sincerely,<br><strong>RTMUN Innovation Portal Outreach Team</strong></p>
                </div>
            </body>
        </html>
        """
        msg_alternative.attach(MIMEText(body_html, "html"))
        
        ics_content = generate_ics_file_content(
            summary=summary,
            description=description,
            date_str=date_str,
            time_str=time_str,
            meet_link=meet_link,
            organizer_email=sender_email,
            attendee_email=lead_email
        )
        
        attachment = MIMEBase('text', 'calendar', method='REQUEST')
        attachment.set_payload(ics_content)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment; filename="invite.ics"')
        attachment.add_header('Content-class', 'urn:content-classes:calendarmessage')
        msg.attach(attachment)
        
        ics_part = MIMEText(ics_content, 'calendar; method=REQUEST')
        ics_part.add_header('Content-class', 'urn:content-classes:calendarmessage')
        msg.attach(ics_part)
        
        port = int(smtp_port)
        if port == 465:
            server = smtplib.SMTP_SSL(smtp_host, port, timeout=10)
        else:
            server = smtplib.SMTP(smtp_host, port, timeout=10)
            server.starttls()
            
        server.login(smtp_user, smtp_pass)
        server.sendmail(sender_email, [lead_email], msg.as_string())
        server.quit()
        print(f"Meeting invitation email with .ics attachment sent successfully to {lead_email}!")
        return True
    except Exception as smtp_err:
        print("Failed to send meeting invitation email via SMTP:", smtp_err)
        return False

oauth_states = {}

@app.get("/api/outreach/authorize")
def google_authorize():
    from google_auth_oauthlib.flow import Flow
    config = get_google_client_config()
    if not config:
        raise HTTPException(
            status_code=400, 
            detail="Google OAuth credentials are not configured in backend/.env file. Please populate GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET."
        )
        
    try:
        flow = Flow.from_client_config(
            config,
            scopes=["https://www.googleapis.com/auth/calendar.events"]
        )
        flow.redirect_uri = config["web"]["redirect_uris"][0]
        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent'
        )
        # Store code_verifier using state as the key
        oauth_states[state] = flow.code_verifier
        return {"authorization_url": authorization_url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate authorization URL: {str(e)}")

@app.get("/api/outreach/oauth2callback")
def google_oauth2callback(code: str, state: Optional[str] = None, error: Optional[str] = None):
    from google_auth_oauthlib.flow import Flow
    if error:
        return HTMLResponse(content=f"<h3>Authorization Error</h3><p>{error}</p>", status_code=400)
        
    config = get_google_client_config()
    if not config:
        return HTMLResponse(content="<h3>Configuration Error</h3><p>OAuth configuration is missing in backend/.env.</p>", status_code=400)
        
    try:
        flow = Flow.from_client_config(
            config,
            scopes=["https://www.googleapis.com/auth/calendar.events"]
        )
        flow.redirect_uri = config["web"]["redirect_uris"][0]
        
        # Retrieve the code_verifier using state
        code_verifier = oauth_states.pop(state, None) if state else None
        flow.fetch_token(code=code, code_verifier=code_verifier)
        
        credentials = flow.credentials
        
        token_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "token.json")
        with open(token_path, "w") as token_file:
            token_file.write(credentials.to_json())
            
        html_content = """
        <html>
            <head><title>Authentication Successful</title></head>
            <body style="font-family: sans-serif; text-align: center; padding-top: 5rem; background: #f8f9ff;">
                <h2 style="color: #10b981;">✓ Google Calendar Authorization Successful!</h2>
                <p>The outreach automation system has been authorized to schedule Google Meet meetings.</p>
                <p style="color: #6b7280; font-size: 0.9rem;">You can close this tab and return to the RTMUN dashboard.</p>
                <script>
                    setTimeout(() => { window.close(); }, 5000);
                </script>
            </body>
        </html>
        """
        return HTMLResponse(content=html_content, status_code=200)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HTMLResponse(content=f"<h3>Token Exchange Failed</h3><p>{str(e)}</p>", status_code=500)

@app.get("/api/outreach/oauth-status")
def get_oauth_status():
    token_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "token.json")
    is_authorized = os.path.exists(token_path)
    
    config = get_google_client_config()
    is_configured = config is not None
    
    return {
        "is_configured": is_configured,
        "is_authorized": is_authorized
    }

class OutreachConfig(BaseModel):
    sync_interval: int

IMAP_SYNC_INTERVAL = 30  # background IMAP check interval in seconds. 0 or negative means manual only.

@app.get("/api/outreach/config")
def get_outreach_config():
    global IMAP_SYNC_INTERVAL
    return {"sync_interval": IMAP_SYNC_INTERVAL}

@app.post("/api/outreach/config")
def update_outreach_config(cfg: OutreachConfig):
    global IMAP_SYNC_INTERVAL
    IMAP_SYNC_INTERVAL = cfg.sync_interval
    print(f"Updated IMAP sync interval to: {IMAP_SYNC_INTERVAL} seconds")
    return {"status": "success", "sync_interval": IMAP_SYNC_INTERVAL}

@app.post("/api/outreach/check-replies")
def trigger_check_replies():
    from datetime import datetime
    replies = check_imap_replies_sync()
    return {"status": "success", "checked_at": datetime.now().isoformat(), "new_replies": replies}

class ScheduleMeetingRequest(BaseModel):
    lead_id: str
    date: str
    time: str

@app.post("/api/outreach/schedule-meeting")
def api_schedule_meeting(req: ScheduleMeetingRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM outreach_leads WHERE id = ?", (req.lead_id,))
    lead = cursor.fetchone()
    if not lead:
        conn.close()
        raise HTTPException(status_code=404, detail="Lead not found")
        
    lead = dict(lead)
    
    import uuid
    summary = f"MOU Collaboration: {lead['incubator_name']}"
    description = f"Introductory discussion regarding strategic cooperation and academic collaboration MoU."
    
    gmeet_info = create_gmeet_event(
        summary=summary,
        description=description,
        date_str=req.date,
        time_str=req.time,
        attendee_email=lead["email"]
    )
    
    calendar_event_id = None
    if gmeet_info:
        meeting_link = gmeet_info["meet_link"]
        calendar_event_id = gmeet_info["calendar_event_id"]
        print(f"Successfully scheduled Google Meet call: {meeting_link}")
    else:
        meeting_link = f"https://meet.google.com/abc-{uuid.uuid4().hex[:4]}-xyz"
        calendar_event_id = f"gcal_{uuid.uuid4().hex[:12]}"
        print(f"OAuth not authorized. Using fallback meet link: {meeting_link}")
        
    meeting_id = f"evt_{uuid.uuid4().hex[:8]}"
    
    cursor.execute("SELECT COUNT(*) FROM scheduled_meetings WHERE lead_id = ?", (req.lead_id,))
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO scheduled_meetings (
                id, lead_id, incubator_name, title, date, time, calendar_event_id, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            meeting_id,
            req.lead_id,
            lead["incubator_name"],
            summary,
            req.date,
            req.time,
            calendar_event_id,
            "Scheduled"
        ))
        
    send_meeting_invite_email(
        lead_name=lead["incubator_name"],
        lead_email=lead["email"],
        date_str=req.date,
        time_str=req.time,
        meet_link=meeting_link
    )
    
    meeting_scheduled_at = f"{req.date} at {req.time}"
    cursor.execute('''
        UPDATE outreach_leads 
        SET status = 'Meeting Scheduled',
            meeting_link = ?,
            meeting_scheduled_at = ?
        WHERE id = ?
    ''', (
        meeting_link,
        meeting_scheduled_at,
        req.lead_id
    ))
    
    conn.commit()
    conn.close()
    
    return {
        "status": "success",
        "meeting_link": meeting_link,
        "meeting_scheduled_at": meeting_scheduled_at
    }

@app.get("/api/outreach/meetings")
def get_outreach_meetings():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT m.*, l.meeting_link 
        FROM scheduled_meetings m
        LEFT JOIN outreach_leads l ON m.lead_id = l.id
    ''')
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

@app.post("/api/outreach/meetings/update-status")
def update_meeting_status(req: UpdateMeetingStatusRequest):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM scheduled_meetings WHERE id = ?", (req.meeting_id,))
    meeting = cursor.fetchone()
    if not meeting:
        conn.close()
        raise HTTPException(status_code=404, detail="Meeting not found")
    cursor.execute("UPDATE scheduled_meetings SET status = ? WHERE id = ?", (req.status, req.meeting_id))
    conn.commit()
    conn.close()
    return {"status": "success", "message": f"Meeting status updated to {req.status}."}

@app.get("/api/outreach/calendar-events")
def get_external_calendar_events():
    import requests
    import os
    
    api_key = os.environ.get("GoogleCalender") or os.environ.get("GoogleCalendar")
    if api_key:
        api_key = api_key.strip().strip('"').strip("'")
        
    if not api_key or "AIzaSy" not in api_key:
        return {"status": "mock", "events": [
            {"summary": "Nagpur University Foundation Day", "date": "2026-08-04"},
            {"summary": "Independence Day Holiday", "date": "2026-08-15"},
            {"summary": "Startup Pitch Competition", "date": "2026-09-10"}
        ]}
        
    try:
        # Fetch from Indian Holidays public calendar using their API key
        calendar_id = "en.indian#holiday@group.v.calendar.google.com"
        url = f"https://www.googleapis.com/calendar/v3/calendars/{calendar_id}/events?key={api_key}&maxResults=10&timeMin=2026-01-01T00:00:00Z"
        res = requests.get(url, timeout=5)
        if res.ok:
            items = res.json().get("items", [])
            events = []
            for item in items:
                summary = item.get("summary")
                start = item.get("start", {})
                date_val = start.get("date") or start.get("dateTime", "").split("T")[0]
                if summary and date_val:
                    events.append({"summary": summary, "date": date_val})
            events.sort(key=lambda x: x["date"])
            return {"status": "real", "events": events[:5]}
    except Exception as e:
        print("Error fetching Google Calendar events:", e)
        
    return {"status": "error_fallback", "events": [
        {"summary": "Nagpur University Foundation Day", "date": "2026-08-04"},
        {"summary": "Independence Day Holiday", "date": "2026-08-15"},
        {"summary": "Startup Pitch Competition", "date": "2026-09-10"}
    ]}

# Start periodic background email reply checker daemon thread
import threading
import time

def start_imap_checking_loop():
    def loop():
        global IMAP_SYNC_INTERVAL
        while True:
            try:
                if IMAP_SYNC_INTERVAL > 0:
                    check_imap_replies_sync()
            except Exception as e:
                print("IMAP background checker loop error:", e)
            
            sleep_time = max(5, IMAP_SYNC_INTERVAL) if IMAP_SYNC_INTERVAL > 0 else 5
            time.sleep(sleep_time)
            
    t = threading.Thread(target=loop, daemon=True)
    t.start()

@app.on_event("startup")
def on_startup():
    start_imap_checking_loop()


